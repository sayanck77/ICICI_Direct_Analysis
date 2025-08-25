
import os
import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from datetime import datetime, date, timedelta, time, timezone
from breeze_connect import BreezeConnect
import pytz


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.common.exceptions import NoSuchElementException, WebDriverException, ElementClickInterceptedException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib.parse
import time as tm

import smtplib
# MIMEMultipart send emails with both text content and attachments.
from email.mime.multipart import MIMEMultipart
# MIMEText for creating body of the email message.
from email.mime.text import MIMEText
# MIMEApplication attaching application-specific data (like CSV files) to email messages.
from email.mime.application import MIMEApplication


global current_price
global df
global s_token


main_excel_file_path = "C:\\Users\\Admin\\My_Projects\\ICICI_Direct_Analysis\\INVESTMENT  - 24-Apr-2025.xlsx"
data_dump_folder = "C:\\Users\\Admin\\My_Projects\\ICICI_Direct_Analysis\\DAILY DATA DUMP"
timestamp_file = "C:\\Users\\Admin\\My_Projects\\ICICI_Direct_Analysis\\Session_Key\\Timestamp.txt"
current_status_folder = "C:\\Users\\Admin\\My_Projects\\ICICI_Direct_Analysis\\CURRENT STATUS"
ta_dump = "C:\\Users\\Admin\\My_Projects\\ICICI_Direct_Analysis\\TECHNICAL ANALYSIS DUMP"


new_table = {"Stock": [], "Date": [], 'Close': [], "S_EMA1": [], "S_EMA2": [], "C_EMA1" : [], "C_EMA2": [], "S_RSI_P": [], "S_RSI_L_R": [], 'C_RSI_V': [],  'P_Pos': [], 'C_Pos': []}
df = pd.DataFrame(new_table)

# --- Configuration ---
LOGIN_URL = "https://api.icicidirect.com/apiuser/home"

# --- IMPORTANT: Replace with your actual username and password ---
USERNAME = "WK178100"
PASSWORD = os.getenv("My_ICICI_Password")

# --- CRITICAL: UPDATE THESE SELECTORS BASED ON YOUR WEBSITE'S HTML ---

# Selector for the initial "LOGIN" button that reveals the form (first page)
INITIAL_LOGIN_BUTTON_SELECTOR_TYPE = By.XPATH
INITIAL_LOGIN_BUTTON_SELECTOR_VALUE = "//a[contains(text(), 'LOGIN')]"

# Selector for the "View Apps" button after OTP/login completion
VIEW_APPS_BUTTON_SELECTOR_TYPE = By.ID
VIEW_APPS_BUTTON_SELECTOR_VALUE = "pills-view-tab"

# Selector for the "Login" link after "View Apps" (opens new tab/window)
NEXT_LINK_AFTER_VIEW_APPS_SELECTOR_TYPE = By.XPATH
NEXT_LINK_AFTER_VIEW_APPS_SELECTOR_VALUE = "//a[@href='https://api.icicidirect.com/apiuser/login?api_key=X4Cx7n2k69498TI15%3d753L0H3%2b4V2474' and contains(text(), 'Login') and @target='_blank']"


# --- Selectors for the SECOND Login Page (the new tab/window) ---
# Username field on the second login page
SECOND_LOGIN_USERNAME_FIELD_SELECTOR_TYPE = By.ID
SECOND_LOGIN_USERNAME_FIELD_SELECTOR_VALUE = "txtuid"

# Password/PIN field on the second login page
SECOND_LOGIN_PASSWORD_FIELD_SELECTOR_TYPE = By.ID
SECOND_LOGIN_PASSWORD_FIELD_SELECTOR_VALUE = "txtPass"

# Terms and Conditions checkbox on the second login page
SECOND_LOGIN_TNC_CHECKBOX_SELECTOR_TYPE = By.ID
SECOND_LOGIN_TNC_CHECKBOX_SELECTOR_VALUE = "chkssTnc"

# FINAL SUBMIT BUTTON on this SECOND Login Page (the one that triggers the second OTP)
SECOND_LOGIN_FINAL_SUBMIT_BUTTON_SELECTOR_TYPE = By.ID
SECOND_LOGIN_FINAL_SUBMIT_BUTTON_SELECTOR_VALUE = "btnSubmit"

# Submit button on the first OTP popup
FIRST_OTP_SUBMIT_BUTTON_SELECTOR_TYPE = By.ID
FIRST_OTP_SUBMIT_BUTTON_SELECTOR_VALUE = "btnsubotp"

# Submit button on the second OTP popup (after submitting the second login form)
SECOND_OTP_SUBMIT_BUTTON_SELECTOR_TYPE = By.ID
SECOND_OTP_SUBMIT_BUTTON_SELECTOR_VALUE = "Button1"

# Path to your WebDriver executable (e.g., chromedriver.exe or geckodriver.exe)
DRIVER_PATH = "C:\\Windows\\System32\\chromedriver.exe"

# Choose your browser: 'chrome' or 'firefox'
BROWSER = 'chrome'

# Max time to wait for elements to be present (in seconds)
WAIT_TIME = 15  # General wait time for most elements

# Dedicated wait times for manual OTP entry (both set to 30 seconds)
FIRST_OTP_MANUAL_ENTRY_WAIT_TIME = 25
SECOND_OTP_MANUAL_ENTRY_WAIT_TIME = 25
# --- End Configuration ---


def setup_isec_login():
    global s_token

    filepath = timestamp_file
    todays_date = ist_date_time()[0] # **Calling ist_date_time, returns str, str**

    if os.path.exists(filepath):

      token, date = extract_token_date() # **Calling extract_token_date, returns str, str**

      if date == todays_date:
        s_token = token
        print("Token already exists.\n")
      else:
        print("Hi. Since the current token is obsolete we need to generate new token. Going there....\n")
        tm.sleep(5) # **Calling tm.sleep, takes int or float**

        s_token = automate_login()

    print(f"The generated token number is {s_token}. \n")

    isec, login_outcome = isec_login(s_token)
    return isec, login_outcome


def ist_date_time():
    # Get the Indian/Maldives timezone object (which covers IST)
    india_timezone = pytz.timezone('Asia/Kolkata')
    # Get the current time in UTC
    utc_now = datetime.now(timezone.utc) # **Calling datetime.now, returns datetime object**
    # Localize the UTC time to the Indian timezone
    ist_now = utc_now.astimezone(india_timezone) # **Calling astimezone, returns datetime object**
    # Return the date part as a string and the time as a string
    return ist_now.date().strftime('%Y-%m-%d'), ist_now.time().strftime("%H:%M:%S") # **Calling date() and time() methods, returns date and time objects, then strftime() returns string**


def extract_token_date():
    # Specify the filename
    file_name = timestamp_file # Make sure this matches the filename used to save
    try:
        # Open the file in read mode ('r')
        with open(file_name, "r") as f:
            # Read the entire content of the file
            file_content = f.read()
        # The content is expected to be in the format "session_key,date"
        # Split the string by the comma
        parts = file_content.split(',')
        # The date is the second part (index 1)
        if len(parts) > 1:
            extracted_token_str = parts[0]
            extracted_date_str = parts[1]
            return extracted_token_str, extracted_date_str
        else:
            print("Could not find the file content to separate the session key and date.")
            sys.exit()
    except FileNotFoundError:
        print(f"Error: The file '{file_name}' was not found.")
        sys.exit()
    except Exception as e:
        print(f"An error occurred during the token date extraction : {e}")
        sys.exit()


def automate_login():
    driver = None
    original_window = None
    api_session_key = None
    try:
        if BROWSER == 'chrome':
            service = ChromeService(executable_path=DRIVER_PATH)
            driver = webdriver.Chrome(service=service)
        elif BROWSER == 'firefox':
            service = FirefoxService(executable_path=DRIVER_PATH)
            driver = webdriver.Firefox(service=service)
        else:
            #print("Unsupported browser specified. Please choose 'chrome' or 'firefox'.")
            return None
        original_window = driver.current_window_handle
        #print(f"Opening {LOGIN_URL} in {BROWSER} browser...")
        driver.get(LOGIN_URL)
        wait = WebDriverWait(driver, WAIT_TIME)
        # --- Step 1: Click the initial LOGIN button (first page) ---
        #print("Looking for the initial LOGIN button...")
        try:
            initial_login_button = wait.until(EC.element_to_be_clickable((INITIAL_LOGIN_BUTTON_SELECTOR_TYPE, INITIAL_LOGIN_BUTTON_SELECTOR_VALUE)))
            initial_login_button.click()
            #print("Initial LOGIN button clicked. Waiting for login form to appear...")
            tm.sleep(3)
        except Exception as e:
            #print(f"Error: Initial LOGIN button not found or not clickable: {e}")
            #print("Please confirm the selector for the initial button that reveals the login form.")
            return None
        # --- Step 2: Fill in Username and Password on the FIRST login form ---
        #print("Attempting to interact with the first login form (if present)...")
        try:
            # Added more specific waits for presence and visibility
            username_field = wait.until(EC.visibility_of_element_located((By.ID, "txtuid")))
            username_field.send_keys(USERNAME)
            #print("Username entered on first form.")
            password_field = wait.until(EC.visibility_of_element_located((By.ID, "txtPass")))
            password_field.send_keys(PASSWORD)
            #print("Password entered on first form.")
            final_login_button = wait.until(EC.element_to_be_clickable((By.ID, "btnlogin")))
            final_login_button.click()
            #print("Final Login button on first form clicked.")
            tm.sleep(5) # Give time for page to process
        except (NoSuchElementException, TimeoutException): # Catching specific exceptions for this optional step
            #print("First login form fields/button not found. Proceeding as if this step is not always present or handled by a modal.")
            pass # Removed 'except Exception as e:' from inside, as it was incorrectly indented
        except Exception as e: # This is the correct placement for a broader catch
            #print(f"Error during first login form interaction: {e}")
            return None
        # --- Pause for manual OTP entry (first time) & Click Submit ---
        #print("\n!!! FIRST OTP POPUP DETECTED !!!")
        #print("Please manually enter the 6-digit OTP sent to your phone in the browser window.")
        #print(f"Pausing script for {FIRST_OTP_MANUAL_ENTRY_WAIT_TIME} seconds to allow for OTP entry.")
        tm.sleep(FIRST_OTP_MANUAL_ENTRY_WAIT_TIME) # Manual OTP entry pause
        #print("Attempting to click 'Submit' button on first OTP popup...")
        try:
            first_otp_submit_button = wait.until(
                EC.element_to_be_clickable((FIRST_OTP_SUBMIT_BUTTON_SELECTOR_TYPE, FIRST_OTP_SUBMIT_BUTTON_SELECTOR_VALUE))
            )
            first_otp_submit_button.click()
            #print("First OTP 'Submit' button clicked.")
            tm.sleep(5)
        except (NoSuchElementException, ElementClickInterceptedException, TimeoutException) as otp1_e:
            #print(f"Error: First OTP Submit button (ID: '{FIRST_OTP_SUBMIT_BUTTON_SELECTOR_VALUE}') not found, not clickable, or timed out.")
            #print(f"Details of first OTP button error: {otp1_e}")
            #print("Please confirm the selector for the FIRST OTP submit button (`btnsubotp`) and that it becomes visible and clickable after OTP entry.")
            return None
        # --- Step 5: Click the "View Apps" button ---
        #print("Looking for 'View Apps' button...")
        try:
            view_apps_button = wait.until(EC.element_to_be_clickable((VIEW_APPS_BUTTON_SELECTOR_TYPE, VIEW_APPS_BUTTON_SELECTOR_VALUE)))
            view_apps_button.click()
            #print("'View Apps' button clicked. Giving more time for subsequent elements to load...")
            tm.sleep(10) # Increased sleep to accommodate potential overlays after View Apps click
        except Exception as e:
            #print(f"Error: 'View Apps' button not found or not clickable: {e}")
            #print("Please confirm the selector for the 'View Apps' button. It might not be visible or clickable immediately after OTP.")
            return None
        # --- Step 6: Click the "Login" link after "View Apps" (opens new tab) ---
        #print("Looking for the 'Login' link after 'View Apps' (which opens a new tab)...")
        try:
            old_window_handles = driver.window_handles
            next_link_button = wait.until(EC.element_to_be_clickable((NEXT_LINK_AFTER_VIEW_APPS_SELECTOR_TYPE, NEXT_LINK_AFTER_VIEW_APPS_SELECTOR_VALUE)))
            # --- Handle ElementClickInterceptedException with JavaScript fallback ---
            try:
                next_link_button.click() # Attempt normal click first
                #print("'Login' link after 'View Apps' clicked.")
            except ElementClickInterceptedException:
                #print("Element click intercepted, trying JavaScript click for 'Login' link...")
                driver.execute_script("arguments[0].click();", next_link_button)
                #print("JavaScript click executed for 'Login' link.")
            # --- End JavaScript fallback ---
            tm.sleep(5) # Give browser time to open new tab/window
            # --- Switch to the new window/tab (using simpler logic) ---
            wait.until(EC.number_of_windows_to_be(2))
            driver.switch_to.window(driver.window_handles[-1])
            #print("Switched to the new login window.")
            tm.sleep(7) # Give new page ample time to load all elements
        except Exception as e:
            #print(f"Error: 'Login' link after 'View Apps' not found or not clickable, or failed to switch window: {e}")
            #print("Please re-inspect the 'Login' link *after* clicking 'View Apps'. The selector might have changed or the element might not be visible/clickable immediately.")
            return None
        # --- Step 7: Fill in Username and Password on the SECOND login page ---
        #print("Looking for username field on SECOND login page...")
        try:
            second_login_username_field = wait.until(EC.presence_of_element_located((SECOND_LOGIN_USERNAME_FIELD_SELECTOR_TYPE, SECOND_LOGIN_USERNAME_FIELD_SELECTOR_VALUE)))
            second_login_username_field.send_keys(USERNAME)
            #print("Username entered on second login page.")
        except Exception as e:
            #print(f"Error: Username field on second login page (ID: '{SECOND_LOGIN_USERNAME_FIELD_SELECTOR_VALUE}') not found or not interactable: {e}")
            #print("Please confirm the selector for the username input field on the second login page.")
            return None
        #print("Looking for password field on SECOND login page...")
        try:
            second_login_password_field = wait.until(EC.presence_of_element_located((SECOND_LOGIN_PASSWORD_FIELD_SELECTOR_TYPE, SECOND_LOGIN_PASSWORD_FIELD_SELECTOR_VALUE)))
            second_login_password_field.send_keys(PASSWORD)
            #print("Password entered on second login page.")
        except Exception as e:
            #print(f"Error: Password field on second login page (ID: '{SECOND_LOGIN_PASSWORD_FIELD_SELECTOR_VALUE}') not found or not interactable: {e}")
            #print("Please confirm the selector for the password input field on the second login page.")
            return None
        # --- Step 8: Check the "I agree to the terms and conditions" box ---
        #print("Looking for T&C checkbox on SECOND login page...")
        try:
            tnc_checkbox = wait.until(EC.element_to_be_clickable((SECOND_LOGIN_TNC_CHECKBOX_SELECTOR_TYPE, SECOND_LOGIN_TNC_CHECKBOX_SELECTOR_VALUE)))
            if not tnc_checkbox.is_selected():
                tnc_checkbox.click()
                #print("T&C checkbox clicked.")
            else:
                #print("T&C checkbox already checked.")
                pass
        except Exception as e:
            #print(f"Error: T&C checkbox on second login page (ID: '{SECOND_LOGIN_TNC_CHECKBOX_SELECTOR_VALUE}') not found or not clickable: {e}")
            #print("Please confirm the selector for the 'I agree to the terms and conditions' checkbox.")
            return None
        # --- Step 9: Click the final submit button on the SECOND login page ---
        #print("Looking for the FINAL SUBMIT button on SECOND login page...")
        try:
            second_login_submit_button = wait.until(EC.element_to_be_clickable((SECOND_LOGIN_FINAL_SUBMIT_BUTTON_SELECTOR_TYPE, SECOND_LOGIN_FINAL_SUBMIT_BUTTON_SELECTOR_VALUE)))
            second_login_submit_button.click()
            #print("FINAL SUBMIT button on second login page clicked.")
            tm.sleep(5) # Give more time for the next OTP popup to appear and load fully
            # --- Pause for manual OTP entry (second time) & Click Submit ---
            #print("\n!!! SECOND OTP POPUP DETECTED !!!")
            #print("Please manually enter the 6-digit OTP sent to your phone for the second login.")
            #print(f"Pausing script for {SECOND_OTP_MANUAL_ENTRY_WAIT_TIME} seconds to allow for OTP entry.")
            tm.sleep(SECOND_OTP_MANUAL_ENTRY_WAIT_TIME) # Manual OTP entry pause
            #print("Looking for 'Submit' button on second OTP popup...")
            try:
                # Increased specific wait time for this critical button
                otp_submit_button = WebDriverWait(driver, WAIT_TIME * 2).until(
                    EC.element_to_be_clickable((SECOND_OTP_SUBMIT_BUTTON_SELECTOR_TYPE, SECOND_OTP_SUBMIT_BUTTON_SELECTOR_VALUE))
                )
                otp_submit_button.click()
                #print("Second OTP 'Submit' button clicked.")
                tm.sleep(7) # Give sufficient time for redirection after OTP submission
            except (NoSuchElementException, ElementClickInterceptedException, TimeoutException) as otp_e:
                #print(f"Error: OTP Submit button (ID: '{SECOND_OTP_SUBMIT_BUTTON_SELECTOR_VALUE}') not found, not clickable, or timed out.")
                #print(f"Details of OTP button error: {otp_e}")
                #print("Please confirm the selector for the OTP submit button (`Button1`) and that it becomes visible and clickable after OTP entry.")
                return None
            # --- Fetch API Session Key ---
            #print("Attempting to fetch API session key from current URL using polling...")
            max_attempts = 20 # Check for 20 seconds
            api_session_key_found = False
            for i in range(max_attempts):
                current_url = driver.current_url
                #print(f"Attempt {i+1}/{max_attempts}: Current URL: {current_url}")
                parsed_url = urllib.parse.urlparse(current_url)
                query_params = urllib.parse.parse_qs(parsed_url.query)
                if 'apisession' in query_params and query_params['apisession']:
                    api_session_key = query_params['apisession'][0]
                    #print(f"\n--- API Session Key Found: {api_session_key} ---\n")
                    api_session_key_found = True
                    break
                tm.sleep(1)
            if not api_session_key_found:
                #print("API Session Key not found in the URL after polling. Final URL observed:")
                #print(current_url)
                api_session_key = None # Ensure it's None if not found after all attempts
        except Exception as e:
            #print(f"An unexpected error occurred during final submit, OTP interaction, or API key fetch: {e}")
            #print(f"Details: {e}")
            return None
        #print("Login and navigation up to second OTP submission completed. Browser will close shortly.")
        tm.sleep(7)
    except WebDriverException as e:
        #print(f"WebDriver error: {e}")
        #print("Please ensure your WebDriver (chromedriver.exe) is correctly installed at the specified path and its version matches your Chrome browser version.")
        #print("If the versions mismatch, you will see errors like 'Session not created: this version of ChromeDriver only supports Chrome version X'.")
        return None
    except Exception as e:
        #print(f"An unexpected error occurred: {e}")
        return None
    finally:
        if driver:
            driver.quit()
            #print("Browser closed.")

    save_sessionkey_date(api_session_key)
    return api_session_key


def save_sessionkey_date(api_session_key):

    # Get the current timestamp
    current_date = ist_date_time()[0]  # **Calling ist_date_time, returns str, str** # Get the date part only

    # Specify the filename
    file_name = timestamp_file # You can change the path and filename

    # Open the file in write mode ('w') - this will create the file if it doesn't exist or overwrite it if it does
    with open(file_name, "w") as f:
      # Combine the session key and timestamp into a single string and write to the file
      f.write(f"{api_session_key},{current_date}")
      print(f"Session key and date saved to {file_name}. \n")
      f.close()


def isec_login(s_token):

    try:
        isec = BreezeConnect(api_key="X4Cx7n2k69498TI15=753L0H3+4V2474")
        isec.generate_session(api_secret="319$5126u_2r37091nz4o51G79a49L+1", session_token = s_token)
        return isec, "Login Success"

    except Exception as e:
        return "Could not authenticate credentials. Please check session key.", "Try Again."


def to_fetch_or_not(isec):

    stock_name_list = fetch_stock_list() # **Calling fetch_stock_list**

    print("Hi, Initiating Historical data fetch \n")
    for stock_name in stock_name_list:
        path = os.path.join(data_dump_folder, f'{stock_name}.csv')

        status, start_date_hist_data = check_hist_data_exist(stock_name) # **Calling check_hist_data_exist, returns str, date object or str**

        if status == "Fresh Download Required":
            print(f"Historical data file for {stock_name} does NOT EXIST. Downloading data...")

            main_data = fetch_hist_data(stock_name, isec, start_date_hist_data) # **Calling fetch_hist_data, takes str, BreezeConnect object, date object or str**

            main_data.to_csv(path, index=False)
            print(f"Historical data fetch done for {stock_name}.\n")

        elif status == "More data fetch required":
            print(f"Historical data file for {stock_name} is NOT UP-to-DATE. Downloading data...")

            addi_stock_data = fetch_hist_data(stock_name, isec, start_date_hist_data) # **Calling fetch_hist_data, takes str, BreezeConnect object, date object or str**

            prev_data = pd.read_csv(path)
            main_data = pd.concat([prev_data, addi_stock_data], ignore_index=True)
            main_data.to_csv(path, index=False)
            print(f"Historical data fetch done for {stock_name}.\n")

        elif status == "Data is up to date":
            print(f"Recent Historical data file for {stock_name} ALREADY EXIST. No need to download fresh data.\n")
            continue

    print("\nCongratulations !! Data Extraction completed for ALL stocks. \n")


def fetch_stock_list():
    file = main_excel_file_path
    sheet = "INVESTMENT"
    inp_df = pd.read_excel(file, sheet_name=sheet)
    inp_df = inp_df.loc[~inp_df["Stock Symbol"].isna()]
    stock_list = list(inp_df["Stock Symbol"])
    return stock_list


def check_hist_data_exist(name_of_stock):
    path = os.path.join(data_dump_folder, f'{name_of_stock}.csv')
    if not os.path.exists(path):
        start_date_hist_data = "1990-01-01" # **str type**
        return "Fresh Download Required", start_date_hist_data

    elif os.path.exists(path):
        data = pd.read_csv(path)

        data = hist_data_transform(path) # **Calling hist_data_transform**

        last_date = data["Date"].iloc[-1] # **date object type**
        start_date_hist_data = last_date + timedelta(days=1) # **Adding timedelta to a date object, returns date object**

        today_date_obj = datetime.strptime(ist_date_time()[0], '%Y-%m-%d').date()

        if start_date_hist_data < today_date_obj: # **Comparing date object with date.today() which returns a date object**
            return "More data fetch required", start_date_hist_data
        else:
            return "Data is up to date", start_date_hist_data


def fetch_hist_data(name_of_stock, isec, start_date_hist_data):
    start_date = str(start_date_hist_data) # **Converting date object to str**

    end_date = get_yesterdays_date() # **Calling get_yesterdays_date, returns str**

    time_interval = "1day"
    exchange_name = "NSE"
    product_cat = "cash"
    data = isec.get_historical_data(interval = time_interval, from_date = start_date, to_date = end_date, stock_code = name_of_stock, exchange_code = exchange_name, product_type = product_cat)
    stock_data = pd.DataFrame(data["Success"])
    if stock_data.empty:
        return stock_data
    else:
        columns_to_drop = ['stock_code', 'exchange_code', 'product_type', 'expiry_date', 'right', 'strike_price', 'open_interest', 'count']
        stock_data.drop(columns=columns_to_drop, inplace=True)
        return stock_data


def get_yesterdays_date():
    date_yesterday = (datetime.strptime(ist_date_time()[0], '%Y-%m-%d').date() - timedelta(days=1)).strftime('%Y-%m-%d') # **Calling ist_date_time, returns str, str. Then converting str to datetime object, subtracting timedelta, and formatting back to str.**
    return date_yesterday # **Returns str**


def hist_data_transform(path):
    data = pd.read_csv(path, parse_dates=['datetime']) # **Using parse_dates to convert column to datetime objects**
    data = data.rename(columns={"datetime": "Date", "close": "Close", "open": "Open", "high": "High", "low": "Low",
                                "volume": "Volume"})
    data['Date'] = pd.to_datetime(data['Date']).dt.date # **Converting to datetime and then extracting date object**
    data = data.sort_values("Date") # **Sorting by date object**
    return data



def delete_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"\nThe file '{filepath}' has been deleted.\n")
    else:
        print(f"\nThe file '{filepath}' does not exist.\n")



def data_handling():

    write_inv_cur_portfio() # **Calling write_inv_cur_portfio**

    fetch_funct_values() # **Calling fetch_funct_values**

    current_status_save_to_excel(df) # **Calling current_status_save_to_excel**

    write_to_inv_D() # **Calling write_to_inv_D**

    write_pos_values_inv() # **Calling write_pos_values_inv**

    send_email() # **Calling send_email**

    print("\n********* CONGRATULATIONS ....... ALL TASKS HAVE BEEN SUCCESSFULLY COMPLETED *********")



def write_inv_cur_portfio():

    filepath = main_excel_file_path
    book = load_workbook(filepath)
    sheet = book['INVESTMENT']

    cur_df = extracting_cur_portfolio_icicidirect() # **Calling extracting_cur_portfolio_icicidirect**

    if "authenticate" in cur_df:
        print(cur_df)

    else:
        print("\nInitiating writing current portfolio details to INV file.... ")
        for index, row in cur_df["stock_code"].items():
            for each_element in range(2, sheet.max_row + 1):
                if sheet.cell(row = each_element, column = 2).value == row:
                    sheet.cell(row = each_element, column = 7).value = int(cur_df.loc[index, "quantity"])
                    sheet.cell(row = each_element, column = 8).value = float(cur_df.loc[index, "average_price"])
                    sheet.cell(row = each_element, column = 9).value = float(cur_df.loc[index, "current_market_price"])
        book.save(filepath)
        print("Completed current portfolio writing on INV file.\n")


def extracting_cur_portfolio_icicidirect():

    data_f = datetime.strptime(ist_date_time()[0], '%Y-%m-%d') - timedelta(2)
    data_from = data_f.isoformat() + ".000Z"
    data_t = datetime.strptime(ist_date_time()[0], '%Y-%m-%d')
    data_till = data_t.isoformat() + ".000Z"

    isec, outcome = isec_login(s_token) # **Calling isec_login**
    if outcome == "Login Success":
        data = isec.get_demat_holdings()
        demat_data = pd.DataFrame(data["Success"])
        data1 = isec.get_portfolio_holdings(exchange_code="NSE", from_date = data_from, to_date = data_till, stock_code="", portfolio_type="") # **Using data_from and data_till (strings)**
        portofolio_data = pd.DataFrame(data1["Success"])

        for i in range(len(portofolio_data)) :
            for j in range(len(demat_data)) :
                if portofolio_data.loc[i, "stock_code"] == demat_data.loc[j, 'stock_code'] :
                    portofolio_data.loc[i, 'stock_ISIN'] = demat_data.loc[j, 'stock_ISIN']
        return portofolio_data.loc[0:, ['stock_code', "stock_ISIN", 'quantity', 'average_price', 'current_market_price']]

    elif "authenticate" in isec:
        return isec + outcome


def fetch_funct_values():
    file = main_excel_file_path
    sheet = "EMA + RSI"
    inp_df = pd.read_excel(file, sheet_name=sheet)
    inp_df = inp_df.loc[~inp_df["Short Name"].isna()]
    sub_df = inp_df[["Short Name", "EMA-Fast", "EMA-Slow", "RSI_Period", "RSI_L_Range"]]
    sub_df = sub_df.reset_index(drop = True)
    print("\nFetching of functional values completed. \n")
    for i in range(0, len(sub_df)):

        cross_over_signal_data(sub_df.loc[i,"Short Name"], int(sub_df.loc[i,"EMA-Fast"]), int(sub_df.loc[i,"EMA-Slow"]), int(sub_df.loc[i,"RSI_Period"]), int(sub_df.loc[i,"RSI_L_Range"])) # **Calling cross_over_signal_data**

    print(tabulate(df[["Stock", "Date", "Close", "S_EMA1", "S_EMA2", "C_EMA1", "C_EMA2", "S_RSI_P", "S_RSI_L_R", 'C_RSI_V', "P_Pos", "C_Pos"]], headers = 'keys', tablefmt = 'psql'))


def cross_over_signal_data(name, EMA1, EMA2, RSI_p, RSI_range):
    print(f"We are working on {name} for signal generation --> ")
    path_to_ta_dump = os.path.join(ta_dump, f'{name}.csv')
    path = os.path.join(data_dump_folder, f'{name}.csv')
    data = pd.read_csv(path)

    data = hist_data_transform(path) # **Calling hist_data_transform**

    # Get current IST date and time
    current_ist_date_str, current_ist_time_str = ist_date_time()

    current_ist_time = datetime.strptime(current_ist_time_str, "%H:%M:%S").time()
    current_ist_day = datetime.strptime(current_ist_date_str, '%Y-%m-%d').strftime('%A')


    if current_ist_day == "Sunday" or current_ist_day == "Saturday": # **Comparing string**
        pass
    elif current_ist_time >= time(9, 30) and current_ist_time <= time(15, 20): # **Comparing time objects**

        cur_val = connect_ws(name) # **Calling connect_ws**

        if cur_val != 0:
            data_1 = {"Date": [datetime.now().date()], "Open": [0], "High": [0], "Low": [0], "Close": [cur_val], "Volume": [0]} # **Getting current date object**
            new_row = pd.DataFrame(data_1)
            data = pd.concat([data, new_row], ignore_index=True)
        else:
            data = fresh_download(name) # **Calling fresh_download**

    data["C_EMA1"] = data.Close.ewm(alpha=(2 / (EMA1 + 1)), min_periods=EMA1).mean()
    data["C_EMA2"] = data.Close.ewm(alpha=(2 / (EMA2 + 1)), min_periods=EMA2).mean()
    data['Signal'] = 0
    data['Signal'] = np.where(data["C_EMA1"] > data["C_EMA2"], 1, 0)
    data["S_RSI_P"] = RSI_p
    # Calculate RSI
    delta = data['Close'].diff()
    gain = (delta.where(delta > 0, 0)).fillna(0)
    loss = (-delta.where(delta < 0, 0)).fillna(0)
    adjust: bool = False
    ignore_na: bool = True
    avg_gain = gain.ewm(alpha=(1 / RSI_p), adjust=adjust, ignore_na=ignore_na, min_periods=2).mean()
    avg_loss = loss.abs().ewm(alpha=(1 / RSI_p), adjust=adjust, ignore_na=ignore_na, min_periods=2).mean()
    # Calculate RS and RSI
    rs = avg_gain / avg_loss
    data['C_RSI_V'] = 100 - (100 / (1 + rs))
    data['C_Pos'] = 0
    data["C_Pos"] = data['Signal'] * data['C_RSI_V']
    data['C_Pos'] = data['C_Pos'].apply(lambda x: 'B' if x < RSI_range and x > 0 else "S" if x == 0 else 'B_N_Y')
    data_table = data[["Date", "Close", "C_EMA1", "C_EMA2", "S_RSI_P", 'C_RSI_V', "C_Pos"]] # **Using Date column (date object)**
    data_table.insert(0, 'Stock', name)
    data_table.insert(3, 'S_EMA1', EMA1)
    data_table.insert(4, 'S_EMA2', EMA2)
    data_table.insert(8, 'S_RSI_L_R', RSI_range)
    data_table.insert(10, 'P_Pos', cur_p_val(name)) # **Calling cur_p_val**

    data_table.to_csv(path_to_ta_dump, index=False)

    pre_final_table = data_table.tail(1)  # this is a Dataframe
    final_Dataframe_table = pre_final_table.reset_index(drop=True)
    #final_tabulate_table = tabulate(final_Dataframe_table[["Stock", "Date", "Close", "S_EMA1", "S_EMA2", "C_EMA1", "C_EMA2", "S_RSI_P", "S_RSI_L_R", 'C_RSI_V', "P_Pos", "C_Pos"]], headers='keys', tablefmt='psql')  # this is a String
    print("Done. \n")
    table_update(final_Dataframe_table) # **Calling table_update**


def connect_ws(stock):

    isec, outcome = isec_login(s_token) # **Calling isec_login**

    if outcome == "Login Success":
        isec.ws_connect()
        isec.on_ticks = on_ticks # **Assigning on_ticks function**
        isec.subscribe_feeds(exchange_code = "NSE", stock_code = stock, product_type = "cash", get_exchange_quotes = True, get_market_depth = False)
        tm.sleep(30) # **Calling tm.sleep, takes int or float**
        isec.unsubscribe_feeds(exchange_code = "NSE", stock_code = stock, product_type = "cash", get_exchange_quotes = True, get_market_depth = False)
        isec.ws_disconnect()
        return current_price
    else:
        return current_price


def on_ticks(ticks):
    global current_price # Declare current_price as global to modify the global variable
    current_price = ticks["last"] # **Assigning value, type depends on the input ticks**


def fresh_download(name):
    isec, outcome = isec_login(s_token) # **Calling isec_login**

    status, start_date_hist_data = check_hist_data_exist(name) # **Calling check_hist_data_exist, returns str, date object or str**

    if status == "Fresh Download Required":
        print(f"Historical data file for {name} does NOT EXIST. Downloading fresh data...")
        output_path = f"{data_dump_folder}\\{stock_name}.csv"

        main_data = fetch_hist_data(stock_name, isec, start_date_hist_data) # **Calling fetch_hist_data, takes str, BreezeConnect object, date object or str**

        main_data.to_csv(output_path, index=False)
        print(f"Congratulations !! Fresh download of data completed for {name}.\n")

        main_data = hist_data_transform(output_path) # **Calling hist_data_transform**

        return main_data


def cur_p_val(name):
    file = main_excel_file_path
    sheet = "INVESTMENT"
    inp_df = pd.read_excel(file, sheet_name=sheet)
    inp_df = inp_df.loc[~inp_df["P_Pos"].isna()]
    sub_df = inp_df[["Stock Symbol", "P_Pos"]]
    sub_df = sub_df.reset_index(drop = True)
    try:
        new_index = sub_df["Stock Symbol"][sub_df["Stock Symbol"] == name].index[0]
        return sub_df.loc[new_index, "P_Pos"] # **Returning value from DataFrame, type depends on the data**
    except IndexError:
        return "No Prev Data" # **Returning string**


def table_update(final_Dataframe_table):
    global df
    df = pd.concat([df, final_Dataframe_table], ignore_index=True) # **Concatenating DataFrames**


def current_status_save_to_excel(df_to_save): # Renamed parameter
    today_date_str = ist_date_time()[0]
    yesterday_date_obj = datetime.strptime(today_date_str, '%Y-%m-%d').date() - timedelta(days=1)
    yesterday_date_str = yesterday_date_obj.strftime('%Y-%m-%d')

    # Check if the dataframe is empty before saving
    if not df_to_save.empty:
        df_to_save.to_excel(os.path.join(current_status_folder, f'Current_status_{yesterday_date_str}.xlsx'), index=False) # Added index=False and fixed pat # **Using date object in filename string formatting**
        print(f"\nCompleted creation of Current Status file - Current_status_{yesterday_date_str}.xlsx.\n") # **Using date object in print statement**
    else:
        print("DataFrame is empty, skipping saving to Excel.")
        sys.exit()


def write_to_inv_D():
    try:
        today_date_str = ist_date_time()[0]
        yesterday_date_obj = datetime.strptime(today_date_str, '%Y-%m-%d').date() - timedelta(days=1)
        yesterday_date_str = yesterday_date_obj.strftime('%Y-%m-%d')

        filepath = main_excel_file_path
        cur_status_filepath = os.path.join(current_status_folder, f'Current_status_{yesterday_date_str}.xlsx') # Fixed path # **Using date object in filename string formatting**
        if os.path.exists(cur_status_filepath):
            try:
                book = load_workbook(filepath)
                sheet = book['INVESTMENT']
            except FileNotFoundError:
                print(f"Error: The main INV file '{filepath}' was not found.\n") # **Using string in print statement**
                sys.exit()
            except KeyError:
                print(f"Error: Sheet 'INVESTMENT' not found in '{filepath}'.\n") # **Using string in print statement**
                sys.exit()
            try:
                df_current_status = pd.read_excel(cur_status_filepath)  # Load the dataframe inside the if block
                df_current_status = df_current_status[["Stock", "C_Pos"]]
            except Exception as e:
                print(f"Error reading the Current status file '{cur_status_filepath}': {e}.\n")
                sys.exit()
            for index, row in df_current_status["Stock"].items():
                for each_element in range(2, sheet.max_row + 1):
                    if sheet.cell(row=each_element, column=2).value == row:
                        sheet.cell(row=each_element, column=4).value = df_current_status.loc[index, "C_Pos"] # **Assigning value from DataFrame, type depends on the data**
            try:
                book.save(filepath)
                print("Completed INV file writing.\n")
            except Exception as e:
                print(f"Error saving the INV file '{filepath}': {e}.\n")
                sys.exit()
        else:
            print(f"The Current status file '{os.path.basename(cur_status_filepath)}' is not generated yet.\n") # More informative message # **Using string in print statement**
            sys.exit()
    except Exception as e:
        print(f"An unexpected error occurred during writing to INV file: {e}.\n") # Catch other potential errors
        sys.exit()

def write_pos_values_inv():
    today_date_str = ist_date_time()[0]
    today_date_obj = datetime.strptime(today_date_str, '%Y-%m-%d').date()
    yesterday_date_obj = today_date_obj - timedelta(days=1)
    yesterday_date_str = yesterday_date_obj.strftime('%Y-%m-%d')
    bef_yes_ter_day_obj = today_date_obj - timedelta(days=2)
    bef_yes_ter_day_str = bef_yes_ter_day_obj.strftime('%Y-%m-%d')


    time_now = datetime.now() # **Getting current datetime object**
    time_now = time_now.strftime("%Y-%m-%d %H:%M") # **Formatting datetime object to string**
    excel_workbook_path_i = main_excel_file_path
    sheet_i = "POSITIONS"
    try:
        inv_file = load_workbook(excel_workbook_path_i)
        inv_p_sheet = inv_file[sheet_i]
        next_column = inv_p_sheet.max_column + 1
        inv_p_sheet.cell(row=1, column=next_column, value=f"{time_now}") # **Using string in cell value**
    except FileNotFoundError:
        print(f"Error: The file '{excel_workbook_path_i}' was not found.\n") # **Using string in print statement**
        sys.exit()
    except KeyError:
        print(f"Error: Sheet '{sheet_i}' not found in '{excel_workbook_path_i}'.\n") # **Using string in print statement**
        sys.exit()
    except Exception as e:
        print(f"An error occurred while opening the INV file: {e}.\n") # **Using string in print statement**
        sys.exit()
    file_c = None
    current_status_filepath_yesterday = os.path.join(current_status_folder, f'Current_status_{yesterday_date_str}.xlsx') # Fixed path # **Using date object in filename string formatting**
    current_status_filepath_befyesterday = os.path.join(current_status_folder, f'Current_status_{bef_yes_ter_day_str}.xlsx') # Fixed path # **Using date object in filename string formatting**
    if os.path.exists(current_status_filepath_yesterday):
        file_c = current_status_filepath_yesterday # **Assigning string value**
    elif os.path.exists(current_status_filepath_befyesterday):
        file_c = current_status_filepath_befyesterday # **Assigning string value**
    else:
        print(f'The file Current_status_"{yesterday_date_str}".xlsx or "{bef_yes_ter_day_str}".xlsx does not exist.\n') # **Using date objects in print statement**
        sys.exit()
    sheet_c = "Sheet1"
    try:
        inp_df_c = pd.read_excel(file_c, sheet_name = sheet_c) # **Reading excel file using string path**
        inp_df_c = inp_df_c[["Stock", "C_Pos"]]
    except FileNotFoundError:
        print(f"Error: The current status file '{file_c}' was not found (should not happen based on previous check, but defensive).\n") # **Using string in print statement**
        sys.exit()
    except KeyError:
        print(f"Error: Sheet '{sheet_c}' not found in '{file_c}'.\n") # **Using string in print statement**
        sys.exit()
    except Exception as e:
        print(f"An error occurred while reading the current status file: {e}.\n") # **Using string in print statement**
        sys.exit()
    for i, v in inp_df_c["Stock"].items():
        for j in range(2, inv_p_sheet.max_row + 1):
            if inv_p_sheet.cell(row = j, column = 1).value ==  v:
                inv_p_sheet.cell(row = j, column = next_column).value = inp_df_c.loc[i, "C_Pos"] # **Assigning value from DataFrame, type depends on the data**
    try:
        inv_file.save(excel_workbook_path_i) # **Saving excel file using string path**
        print("Added position values in POSITION tab in INV file.\n") # **Using string in print statement**
    except Exception as e:
        print(f"An error occurred while saving the INV file: {e}.\n") # **Using string in print statement**
        sys.exit()


def send_email():
    today_date_str = ist_date_time()[0]
    yesterday_date_obj = datetime.strptime(today_date_str, '%Y-%m-%d').date() - timedelta(days=1)
    yesterday_date_str = yesterday_date_obj.strftime('%Y-%m-%d')

    subject = f"INVESTMENT file sent on {today_date_str}" # Formatted date # **Formatting datetime object to string**
    body = "Check file" # **String type**
    sender_email = "sayanck77@gmail.com" # **String type**
    recipient_email = "sayanck77@gmail.com" # **String type**
    sender_password = "qioy lvvb kmlr deym" # **String type**
    smtp_server = 'smtp.gmail.com' # **String type**
    smtp_port = 465 # **Integer type**
    path_to_file_1 = main_excel_file_path # **String type**
    path_to_file_2 = os.path.join(current_status_folder, f'Current_status_{yesterday_date_str}.xlsx') # Fixed path # **Using date object in filename string formatting**
    # MIMEMultipart() creates a container for an email message that can hold
    # different parts, like text and attachments and in next line we are
    # attaching different parts to email container like subject and others.
    message = MIMEMultipart() # **Creating MIMEMultipart object**
    message['Subject'] = subject # **Assigning string to dictionary key**
    message['From'] = sender_email # **Assigning string to dictionary key**
    message['To'] = recipient_email # **Assigning string to dictionary key**
    body_part = MIMEText(body) # **Creating MIMEText object with string**
    message.attach(body_part) # **Attaching MIMEText object**
    # section 1 to attach file
    try:
        with open(path_to_file_1,'rb') as file: # **Opening file using string path**
            # Attach the file with filename to the email
            message.attach(MIMEApplication(file.read(), Name = os.path.basename(path_to_file_1))) # Use basename for attachment name # **Creating MIMEApplication object and attaching**
    except FileNotFoundError:
        print(f"Error: Attachment file not found: {path_to_file_1}.\n") # **Using string in print statement**
        sys.exit()
    except Exception as e:
        print(f"Error attaching file {path_to_file_1}: {e}.\n") # **Using string in print statement**
        sys.exit()
    try:
        with open(path_to_file_2,'rb') as file: # **Opening file using string path**
            # Attach the file with filename to the email
            message.attach(MIMEApplication(file.read(), Name = os.path.basename(path_to_file_2))) # Use basename for attachment name # **Creating MIMEApplication object and attaching**
    except FileNotFoundError:
        print(f"Error: Attachment file not found: {path_to_file_2}.\n") # **Using string in print statement**
        sys.exit()
    except Exception as e:
        print(f"Error attaching file {path_to_file_2}: {e}.\n") # **Using string in print statement**
        sys.exit()
    # secction 2 for sending email
    try:
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server: # **Connecting to SMTP server using string and integer**
            server.login(sender_email, sender_password) # **Logging in using strings**
            server.sendmail(sender_email, recipient_email, message.as_string()) # **Sending email using strings and message as string**
        print("Email sent to your gmail account - INV file.\n") # **Using string in print statement**
    except Exception as e:
        print(f"Error sending email: {e}.\n") # **Using string in print statement**
        sys.exit()


def clear_df(df):
    df.drop(df.index, inplace=True) # **Modifying DataFrame**
    print("Global DataFrame 'df' cleared.\n") # **Using string in print statement**


################ Main execution block ################

isec, login_outcome = setup_isec_login() #done # **Calling setup_isec_login**

if login_outcome == "Login Success":

    to_fetch_or_not(isec) #done # **Calling to_fetch_or_not**

    today_date_str = ist_date_time()[0]
    yesterday_date_obj = datetime.strptime(today_date_str, '%Y-%m-%d').date() - timedelta(days=1)
    yesterday_date_str = yesterday_date_obj.strftime('%Y-%m-%d')

    delete_file(os.path.join(current_status_folder, f'Current_status_{yesterday_date_str}.xlsx')) #done # **Calling delete_file with string path**

    data_handling() # **Calling data_handling**

    clear_df(df) # **Calling clear_df**

else:

    print(f"Script stopped due to login failure: {isec}") # **Using string in print statement**




